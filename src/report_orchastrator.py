import json
import os
from dataclasses import asdict
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from planning_center_client import PlanningCenterClient
from attendance_decline_accumulator import AttendanceDeclineAccumulator
from report_models import MemberAttendance

class ReportOrchastrator:
    def __init__(self):
        httpClient = PlanningCenterClient()
        self._accumulator = AttendanceDeclineAccumulator(httpClient)

    def generate_report(self):
        group_name = os.getenv("GROUP_NAME")
        comparison_size_weeks = int(os.getenv("COMPARISON_SIZE_WEEKS", "26"))
        decline_threshold = float(os.getenv("DECLINE_THRESHOLD", ".20"))

        EASTERN = ZoneInfo("America/New_York")

        now_eastern = datetime.now(EASTERN)
        days_until_saturday = (5 - now_eastern.weekday()) % 7
        saturday_date = (now_eastern + timedelta(days=days_until_saturday)).date()
        end_date = datetime(saturday_date.year, saturday_date.month, saturday_date.day, tzinfo=EASTERN)
        middle_date = end_date - timedelta(days=comparison_size_weeks * 7 - 1)
        start_date = middle_date - timedelta(days=comparison_size_weeks * 7)

        attendance_report = self._accumulator.get_members_with_declining_attendance(
            group_name, start_date, middle_date, end_date, decline_threshold
        )

        if attendance_report.error_message==None:
            self._write_attendance_report_to_excel(attendance_report.members, start_date, middle_date, "test_output/attendance_report.xlsx")
        else:
            print("Could not generate report: " + attendance_report.error_message)

    def _write_attendance_report_to_excel(self, members: list[MemberAttendance], start_date: datetime, middle_date: datetime, path: str) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Declining Attendance"
        TITLE_ROW = 1
        HEADER_ROW = 3
        TITLE_FONT_SIZE = 16

        headers = [
            "First Name",
            "Last Name",
            "Early Period Attendance",
            "Worship Day Count",
            "Early Period Frequency",
            "Late Period Attendance",
            "Worship Day Count",
            "Late Period Frequency",
            "Frequency Change",
        ]        

        title = f"Attendance Comparison between Period Starting {start_date.date().isoformat()} and Period Starting {middle_date.date().isoformat()}"
        title_cell = sheet.cell(row=TITLE_ROW, column=1, value=title)
        sheet.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=len(headers))
        title_cell.alignment = Alignment(horizontal="center")
        # No name= set here, matching Font(bold=True) below (also no name=), so both inherit the workbook's theme font.
        title_cell.font = Font(size=TITLE_FONT_SIZE)

        PERCENTAGE_FORMAT = "0%"
        PERCENTAGE_COLUMN_HEADERS = ["Early Period Frequency", "Late Period Frequency", "Frequency Change"]
        percentage_columns = [get_column_letter(headers.index(header) + 1) for header in PERCENTAGE_COLUMN_HEADERS]

        sheet.append([])
        sheet.append(headers)
        sheet.row_dimensions[HEADER_ROW].height = 30
        for column_index, cell in enumerate(sheet[HEADER_ROW], start=1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)
            sheet.column_dimensions[cell.column_letter].width = 16

        for member in members:
            row_index = sheet.max_row + 1
            sheet.append([
                member.first_name,
                member.last_name,
                member.early_period_attendance,
                member.early_period_record_count,
                member.early_period_frequency(),
                member.late_period_attendance,
                member.late_period_record_count,
                member.late_period_frequency(),
                member.frequency_change(),
            ])
            for column_letter in percentage_columns:
                sheet[f"{column_letter}{row_index}"].number_format = PERCENTAGE_FORMAT

        workbook.save(path)